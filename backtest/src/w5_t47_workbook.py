#!/usr/bin/env python3
"""
Turn the t47 9-day trade export into a workbook you can check against a chart.

Two views of the same data:
  * "Positions" — one row per trade: ticker, side, entry time/price, initial
    stop, then each scale-out (date, price, realized R) and the final exit.
  * "Legs" — one row per closing leg, the raw record behind the positions view.

Realized R is computed from the position's own entry and initial stop, not from
any label the engine wrote, so TP1 should land near +0.65R, TP2 near +1.85R,
TP3 near +2.75R and a stop near -1R. Anything far off those is worth a look.

Run:  uv run python3 backtest/src/w5_t47_workbook.py
"""
import csv, importlib.util
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
_w = importlib.util.spec_from_file_location("w5", str(HERE / "w5_common.py"))
w5 = importlib.util.module_from_spec(_w); _w.loader.exec_module(w5)

REPORTS = w5.DOE_DIR / "reports"
SRC = REPORTS / "t47_9day_challenge_trades.csv"
DST = REPORTS / "t47_9day_challenge.xlsx"

HDR = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="1F3864")
MONEY = '#,##0.00'


def style(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.font = HDR; c.fill = FILL; c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def main():
    rows = list(csv.DictReader(open(SRC)))
    for r in rows:
        r["counted"] = r["counted_toward_pass"] == "True"
        r["leg"] = int(r["leg"]); r["step"] = int(r["step"])
        for k in ("entry_price", "initial_sl", "exit_price", "pnl", "swap"):
            r[k] = float(r[k] or 0)
        r["realized_R"] = float(r["realized_R"]) if r["realized_R"] else None

    wb = Workbook(); wb.remove(wb.active)

    # ── Summary ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.append(["t47 — fastest challenge pass, 9 calendar days total"])
    ws["A1"].font = Font(bold=True, size=13); ws.append([])
    ws.append(["Step", "Starts", "Passed on day", "Target", "Realized by pass day",
               "Positions", "Legs"])
    for step, start, pd_, tgt in ((1, "2016-05-31", 5, 8000.0), (2, "2016-06-06", 3, 5000.0)):
        sel = [r for r in rows if r["step"] == step and r["counted"]]
        ws.append([step, start, pd_, tgt,
                   round(sum(r["pnl"] + r["swap"] for r in sel), 2),
                   len({r["position"] for r in sel}), len(sel)])
    ws.append([])
    ws.append(["Config", "risk 2.5%/trade · max 20 positions · corr-group cap 6 · "
               "cum-risk 7.5 · TP 0.65R/1.85R/2.75R closing 25%/60%/15%"])
    ws.append(["Account", "fresh $100,000 per step · 5% daily wall · closed-balance targets"])
    ws.append(["Note", "Step 2 starts the day after step 1 passes, on a fresh $100k, "
                       "which is how the 5%ers two-step challenge is scored here."])
    for c in ("A3", "B3", "C3", "D3", "E3", "F3", "G3"):
        ws[c].font = HDR; ws[c].fill = FILL
    for col, w in zip("ABCDEFG", (10, 14, 15, 12, 22, 11, 8)):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=4, max_row=5, min_col=4, max_col=5):
        for c in row: c.number_format = MONEY

    # ── Positions ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Positions")
    ws.append(["Step", "Ticker", "Side", "Entry time", "Entry price", "Initial SL",
               "Risk (price)", "Exit 1 date", "Exit 1 price", "R1",
               "Exit 2 date", "Exit 2 price", "R2",
               "Exit 3 date", "Exit 3 price", "R3",
               "Final exit date", "Final price", "Final R",
               "Legs", "Total P&L", "Counted toward pass"])
    by = defaultdict(list)
    for r in rows:
        by[(r["step"], r["position"])].append(r)
    for (step, pos), legs in sorted(by.items(), key=lambda kv: kv[1][0]["entry_time"]):
        legs.sort(key=lambda r: r["leg"])
        f = legs[0]
        cells = [step, f["symbol"], f["side"], f["entry_time"], f["entry_price"],
                 f["initial_sl"], round(abs(f["entry_price"] - f["initial_sl"]), 6)]
        for i in range(3):
            if i < len(legs) - 1:
                l = legs[i]; cells += [l["exit_time"], l["exit_price"], l["realized_R"]]
            else:
                cells += ["", "", ""]
        last = legs[-1]
        cells += [last["exit_time"], last["exit_price"], last["realized_R"], len(legs),
                  round(sum(l["pnl"] + l["swap"] for l in legs), 2),
                  "YES" if any(l["counted"] for l in legs) else "no"]
        ws.append(cells)
    style(ws, (5, 10, 6, 20, 12, 12, 12, 20, 12, 7, 20, 12, 7, 20, 12, 7, 20, 12, 8, 6, 11, 18))
    for row in ws.iter_rows(min_row=2, min_col=21, max_col=21):
        for c in row: c.number_format = MONEY

    # ── Legs ─────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Legs")
    cols = ["step", "position", "leg", "symbol", "side", "volume", "entry_time",
            "entry_price", "initial_sl", "exit_time", "exit_price", "realized_R",
            "exit_kind", "pnl", "swap", "day_from_start", "counted_toward_pass",
            "mfe_r", "mae_r"]
    ws.append([c.replace("_", " ") for c in cols])
    for r in sorted(rows, key=lambda r: (r["step"], r["entry_time"], r["leg"])):
        ws.append([r.get(c) for c in cols])
    style(ws, (5, 22, 5, 10, 6, 8, 20, 12, 12, 20, 12, 8, 12, 10, 8, 8, 10, 8, 8))

    wb.save(DST)
    print(f"wrote {DST}  ({len(by)} positions, {len(rows)} legs)")


if __name__ == "__main__":
    main()
